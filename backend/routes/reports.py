import io
import logging
from flask import Blueprint, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from models.bom import BOM
from models.product import Product

# 配置日志
logger = logging.getLogger(__name__)

reports_bp = Blueprint('reports', __name__)

@reports_bp.route('/reports/bom/export', methods=['GET'])
def export_bom_excel():
    """导出BOM数据为Excel格式"""
    try:
        # 获取所有产品的BOM信息
        products = Product.get_all()
        bom_data = []
        
        for product in products:
            # 获取展开的BOM
            bom_items = BOM.get_by_product_with_components(product['id'])
            
            # 只有当产品有BOM项时才添加到导出数据中
            if bom_items:
                for item in bom_items:
                    bom_data.append({
                        '产品SKU': product['sku'],
                        '产品名称': product['name'],
                        '物料SKU': item['material_sku'],
                        '物料名称': item['material_name'],
                        '所需数量': float(item['quantity_required']),
                        '单位': item.get('unit', '个'),
                        '单价': float(item.get('material_price', 0)),
                        '小计': float(item.get('item_cost', 0))
                    })
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "BOM报表"
        
        # 设置标题样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 添加表头
        headers = ['产品SKU', '产品名称', '物料SKU', '物料名称', '所需数量', '单位', '单价', '小计']
        ws.append(headers)
        
        # 设置标题行样式
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 添加数据行
        for row_data in bom_data:
            ws.append([
                row_data['产品SKU'],
                row_data['产品名称'],
                row_data['物料SKU'],
                row_data['物料名称'],
                row_data['所需数量'],
                row_data['单位'],
                row_data['单价'],
                row_data['小计']
            ])
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='bom_report.xlsx'
        )
    except Exception as e:
        logger.error("导出BOM Excel报表时出错: %s", str(e), exc_info=True)
        return jsonify({'error': '导出BOM报表失败'}), 500

@reports_bp.route('/reports/material-requirements', methods=['GET'])
def get_material_requirements():
    """生成物料需求计划报表"""
    try:
        # 获取所有产品的BOM信息
        products = Product.get_all()
        material_requirements = {}
        
        for product in products:
            # 获取展开的BOM
            bom_items = BOM.get_by_product_with_components(product['id'])
            
            # 统计物料需求
            for item in bom_items:
                material_id = item['material_id']
                material_sku = item['material_sku']
                material_name = item['material_name']
                required_quantity = float(item['quantity_required'])
                unit = item.get('unit', '个')
                
                if material_id not in material_requirements:
                    material_requirements[material_id] = {
                        '物料SKU': material_sku,
                        '物料名称': material_name,
                        '单位': unit,
                        '总需求数量': 0,
                        '当前库存': float(item.get('current_stock', 0)),
                        '缺货数量': 0
                    }
                
                material_requirements[material_id]['总需求数量'] += required_quantity
                material_requirements[material_id]['缺货数量'] = max(
                    material_requirements[material_id]['总需求数量'] - 
                    material_requirements[material_id]['当前库存'], 
                    0
                )
        
        # 转换为列表并排序
        result = list(material_requirements.values())
        result.sort(key=lambda x: x['缺货数量'], reverse=True)
        
        return jsonify(result)
    except Exception as e:
        logger.error("生成物料需求计划报表时出错: %s", str(e), exc_info=True)
        return jsonify({'error': '生成物料需求计划报表失败'}), 500

@reports_bp.route('/reports/cost-analysis', methods=['GET'])
def get_cost_analysis():
    """显示成本分析报表"""
    try:
        products = Product.get_all()
        cost_data = []
        
        for product in products:
            # 获取展开的BOM
            bom_items = BOM.get_by_product_with_components(product['id'])
            
            # 计算总成本
            total_cost = sum(float(item.get('item_cost', 0)) for item in bom_items)
            
            # 添加到成本分析数据
            cost_data.append({
                '产品SKU': product['sku'],
                '产品名称': product['name'],
                '物料数量': len(bom_items),
                '总成本': float(total_cost),
                '单位成本': float(total_cost) if product.get('quantity', 0) > 0 else 0
            })
        
        # 按总成本排序
        cost_data.sort(key=lambda x: x['总成本'], reverse=True)
        
        return jsonify(cost_data)
    except Exception as e:
        logger.error("生成成本分析报表时出错: %s", str(e), exc_info=True)
        return jsonify({'error': '生成成本分析报表失败'}), 500

@reports_bp.route('/reports/purchase-list', methods=['GET'])
def get_purchase_list():
    """生成采购清单报表"""
    try:
        # 获取所有产品的BOM信息
        products = Product.get_all()
        material_requirements = {}
        
        for product in products:
            # 获取展开的BOM
            bom_items = BOM.get_by_product_with_components(product['id'])
            
            # 统计物料需求
            for item in bom_items:
                material_id = item['material_id']
                material_sku = item['material_sku']
                material_name = item['material_name']
                required_quantity = float(item['quantity_required'])
                unit = item.get('unit', '个')
                current_stock = float(item.get('current_stock', 0))
                material_price = float(item.get('material_price', 0))
                
                if material_id not in material_requirements:
                    material_requirements[material_id] = {
                        '物料SKU': material_sku,
                        '物料名称': material_name,
                        '单位': unit,
                        '总需求数量': 0,
                        '当前库存': current_stock,
                        '采购单价': material_price,
                        '缺货数量': 0,
                        '采购金额': 0
                    }
                
                material_requirements[material_id]['总需求数量'] += required_quantity
                material_requirements[material_id]['缺货数量'] = max(
                    material_requirements[material_id]['总需求数量'] - 
                    material_requirements[material_id]['当前库存'], 
                    0
                )
                material_requirements[material_id]['采购金额'] = (
                    material_requirements[material_id]['缺货数量'] * 
                    material_requirements[material_id]['采购单价']
                )
        
        # 过滤出需要采购的物料（缺货数量 > 0）
        purchase_list = [
            item for item in material_requirements.values() 
            if item['缺货数量'] > 0
        ]
        
        # 按采购金额排序
        purchase_list.sort(key=lambda x: x['采购金额'], reverse=True)
        
        return jsonify(purchase_list)
    except Exception as e:
        logger.error("生成采购清单报表时出错: %s", str(e), exc_info=True)
        return jsonify({'error': '生成采购清单报表失败'}), 500